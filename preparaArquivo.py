from functions import *
import pandas as pd
import openpyxl
import math
from decimal import Decimal

# O OBJETIVO DESTE CÓDIGO É MAPEAR PARA UMA NOVA LISTA APENAS OS PACIENTES QUE REALMENTE PRECISAM RENOVAR,
# USANDO COMO BASE um arquivo apacsVencidas.xlsx E ASSIM CRIANDO UM NOVO ARQUIVO
# COM SOMENTE OS PACIENTES DE apacsVencidas.xlsx e as colunas "medico, inicio, alb
# hb, urr, acesso" que estão em baseDados.xlsx

nome = []
cpf = []
medico = []
inicio = []
alb = []
hb = []
urr = []
acesso = []

dfRenovar = pd.read_excel("apacsVencidas.xlsx")
# display(dfRenovar)

# neste arquivo contem a lista dos pacientes com apacs vencidas
# gerado pelo sistema em Faturamento/SUS/Laudos de Apacs
# com os critérios: Mês vigente/Tratamentos/Sem laudo

dfDados = pd.read_excel("baseDados.xlsx")
# display(dfDados)

# neste arquivo contem os nomes de todos os pacientes ativos com dados 
# de cpf, medico, inicio, alb, hb, urr, acesso gerados pelo relatório 
# do sistema da seguinte forma:
# menu: Clientes/Relatorio de Clientes
# Critérios:
# Período => De:(3 meses pra trás) Até: Dia atual
# Programa: Hemodiálise, Hemodiálise HIV/HB/HC
# Relatório: RENOVAÇÃO APACS
# Somente pacientes com dados
# Somente clientes ativos hoje
# Tipo de resultado: Média

# MANIPULANDO OS ARQUIVOS DO EXCEL =>

# Carrregando arquivo com os pacientes que realmente precisam renovar
book2 = openpyxl.load_workbook('apacsVencidas.xlsx')
# Selecionando uma página
pacientes_real = book2['Sheet']

# Carregando arquivo com a lista que contém os dados de todos os pacientes
book = openpyxl.load_workbook('baseDados.xlsx')
# Selecionando uma página
pacientes_dados = book['Sheet']

# fazendo a filtragem...

# for para o qual quero filtrar:
for rows in pacientes_real.iter_rows(min_row=2):
  nomePaciente = rows[0].value

  # for no qual eu tenho todos os dados de todos os pacientes
  for rows in pacientes_dados.iter_rows(min_row=2):
    if rows[0].value == nomePaciente:
      #convertendo a data para nao aparecer a hora junto:
      data_inicio = f"{rows[3].value:%d/%m/%Y}"
      # populando os arrays dos campos separadamente:
      nome.append(rows[0].value)
      cpf.append(rows[1].value)
      medico.append(rows[2].value.title())
      inicio.append(data_inicio)
      alb.append(str(math.trunc(Decimal(rows[4].value.replace(',','.'))*10)))
      hb.append(str(math.trunc(Decimal(rows[5].value.replace(',','.'))*10)))
      urr.append(str(math.trunc(Decimal(rows[6].value.replace(',','.'))*100)))
      acesso.append(rows[7].value)
# fim do processo com ambos arquivos do excel iniciais

# montando o json que ira servir de base para re-converter em excel:
json_file = {
      'Nome':nome,
      'CPF':cpf,
      'Medico':medico,
      'Inicio':inicio,
      'Alb':alb,
      'Hb':hb,
      'Urr':urr,
      'Acesso':acesso
      }

# convertendo o objeto json em um DataFrame(planilha)
df = pd.DataFrame(json_file)
# display(df) 

# salvando o resultado para uma nova planilha de excel
df.to_excel("ApacsRenovar.xlsx", index=None, sheet_name="Sheet")

# - mudar os nomes para os mesmos como aparecem no trs

# O OBJETIVO DESTE CÓDIGO ABAIXO É PREPARAR O ARQUIVO FINAL PRA RENOVAÇÃO
# PEGANDO OS NOMES REAIS DOS PACIENTES EXATAMENTE COMO ESTÃO NO TRS
# PORQUE A PARTIR DOS NOMES REAIS OUTRO CÓDIGO SOLICITARÁ AS RENOVAÇÕES (renovarApacs.py)

# MANIPULANDO OS ARQUIVOS DO EXCEL
# Carregando arquivo com a lista de apacs para renovar
book = openpyxl.load_workbook('ApacsRenovar.xlsx')
# Selecionando uma página
pacientes_renova = book['Sheet']

for rows in pacientes_renova.iter_rows(min_row=2):
    cpf = rows[1].value # cpf na planilha de apacs vencidas
    nomeTrs = acharCpf(cpf) # Pega na Planilha 'PacientesEmTratamento v2' os nomes exatamente como estão no TRS
    rows[0].value = nomeTrs # Atribui esses nomes na planilha ApacsRenovar pra solicitar renovação pelo nome
    
    # Colocar a primeira letra em maiuscula e o restante em minuscula de cada parte do nome do medico
    rows[2].value = (rows[2]).value.title()

book.save('ApacsRenovar.xlsx')